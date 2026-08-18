import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def generate_production_excel():
    excel_path = "Production_Breakout_Master_V3.xlsx"
    print(f"Building Production Master Excel Report: '{excel_path}'...")

    df_5y = pd.read_csv("recent_ipo_breakouts_5y.csv") if os.path.exists("recent_ipo_breakouts_5y.csv") else pd.DataFrame()
    df_daily = pd.read_csv("master_scan_results.csv") if os.path.exists("master_scan_results.csv") else pd.DataFrame()
    df_monthly = pd.read_csv("monthly_ath_breakouts.csv") if os.path.exists("monthly_ath_breakouts.csv") else pd.DataFrame()

    df_spec = pd.DataFrame([
        {"Specification": "Quantitative Architecture", "Value": "Enterprise Production Engine V3.0"},
        {"Specification": "Bollinger Bands Setting", "Value": "Period = 20, StdDev = 2.0 (Upper Band Contraction & Expansion)"},
        {"Specification": "Fast Momentum RSI", "Value": "Period = 9, Threshold > 60 (Acceleration Filter)"},
        {"Specification": "Volume Filter", "Value": "Volume >= 1.5x 20-day Average Volume & Z-Score Analysis"},
        {"Specification": "Stop Loss (SL)", "Value": "Dynamic ATR Trailing SL (2x ATR) / 20 SMA Support Zone"},
        {"Specification": "Target 1 (Conservative)", "Value": "Entry + 1.5x Risk (Fibonacci 1.272 ATH Extension)"},
        {"Specification": "Target 2 (Aggressive)", "Value": "Entry + 2.5x Risk (Fibonacci 1.618 ATH Extension)"},
        {"Specification": "Target 3 (Institutional Moonshot)", "Value": "Entry + 4.0x Risk (Fibonacci 2.618 ATH Extension)"},
        {"Specification": "Scanned Stock Universes", "Value": "BSE Main Board IPOs (2021-2026) + NSE Main Board IPOs (2021-2026) + NSE Top Stocks"}
    ])

    try:
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            if not df_5y.empty:
                df_5y.to_excel(writer, sheet_name='5-Year IPO Breakouts', index=False)
            if not df_daily.empty:
                df_daily.to_excel(writer, sheet_name='Daily Institutional Breakouts', index=False)
            if not df_monthly.empty:
                df_monthly.to_excel(writer, sheet_name='Monthly ATH Breakouts', index=False)
            df_spec.to_excel(writer, sheet_name='Quantitative System Protocol', index=False)

        wb = openpyxl.load_workbook(excel_path)
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Arial", size=10)

        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            ws.views.sheetView[0].showGridLines = True

            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = data_font
                    cell.border = thin_border
                    val_str = str(cell.value)
                    if val_str.replace('.', '', 1).isdigit() or val_str.startswith('+') or val_str.endswith('%'):
                        cell.alignment = Alignment(horizontal="center", vertical="center")

            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(excel_path)
        print(f"SUCCESS: Production Master Excel report saved at '{excel_path}'!")
    except Exception as e:
        print(f"[Excel Error] {e}")

if __name__ == "__main__":
    generate_production_excel()
